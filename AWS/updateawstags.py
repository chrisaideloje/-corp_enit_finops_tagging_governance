import os

import boto3

import openpyxl

import logging
import argparse



# Setup logging configuration

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')



# Load AWS credentials

#id = credentials.AWS_ACCESS_KEY_ID

id = os.getenv("AWS_ACCESS_KEY_ID")

key = os.getenv("AWS_SECRET_ACCESS_KEY")

region = os.getenv("AWS_DEFAULT_REGION")



def create_aws_session():

    """

    Create an AWS session using credentials loaded from the credentials module.

    """

    # Create an STS client using permanent credentials

    sts_client = boto3.client(

        'sts',

        aws_access_key_id=id,

        aws_secret_access_key=key,

        region_name=region

    )



    # Get temporary security credentials

    response = sts_client.get_session_token(DurationSeconds=3600)



    # Access temporary credentials from the response

    temp_credentials = response['Credentials']



    return boto3.Session(

        aws_access_key_id=temp_credentials['AccessKeyId'],

        aws_secret_access_key=temp_credentials['SecretAccessKey'],

        aws_session_token=temp_credentials['SessionToken'],

        region_name=region

    )



def load_excel_file(file_path):

    """

    Load the Excel file and return account IDs.



    :param file_path: The path to the Excel file

    :return: List of account IDs

    """

    account_ids = []

    try:

        workbook = openpyxl.load_workbook(file_path)

        sheet = workbook.active

       

        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming account IDs start from the second row

            account_id = row[0]  # Assuming the account ID is in the first column

            if account_id is not None:

                account_ids.append(account_id)  # Retain original format

           

    except Exception as e:

        logging.error(f"Failed to load Excel file {file_path}: {e}")



    return account_ids



def list_tags(account_id, org_client):

    """

    List tags for a given AWS Organizations account.



    :param account_id: The ID of the AWS account

    :param org_client: AWS Organizations client

    :return: Tags associated with the account

    """

    try:

        response = org_client.list_tags_for_resource(

            ResourceId=account_id

        )

        return response['Tags']

    except Exception as e:

        logging.error(f'Error retrieving tags for account {account_id}: {str(e)}')

        return None



def write_tags_to_excel(output_file, account_tag_map):

    """

    Write account IDs and their corresponding tags to a new Excel file.



    :param output_file: The path for the output Excel file

    :param account_tag_map: A dictionary mapping account IDs to their tags

    """

    workbook = openpyxl.Workbook()

    sheet = workbook.active

   

    # Write headers

    headers = ['Account ID']

   

    # Create a set to hold unique tag keys

    unique_keys = set()

    for tags in account_tag_map.values():

        if tags:

            for tag in tags:

                unique_keys.add(tag['Key'])

   

    headers += list(unique_keys)

    sheet.append(headers)

   

    # Write account IDs and tags

    for account_id, tags in account_tag_map.items():

        row = [account_id]  # Account ID as is

        tag_dict = {tag['Key']: tag['Value'] for tag in tags} if tags else {}

       

        # Add values for each unique key; if the key is not present, add an empty string

        for key in unique_keys:

            row.append(tag_dict.get(key, ''))

       

        sheet.append(row)



    # Save the workbook

    try:

        workbook.save(output_file)

        logging.info(f"Tags successfully written to {output_file}.")

    except Exception as e:

        logging.error(f"Failed to write tags to Excel file {output_file}: {e}")



def main():

    excel_file = "C:\\Users\\H630384\\project\\corp_enit_finops_tagging_governance\\AWS\\AWS_Account_Ids.xlsx" # Replace with your actual Excel file path

    output_file = "AWS_Account_Tags_New.xlsx"  # Output file name

    account_ids = load_excel_file(excel_file)



    if not account_ids:

        logging.info("No account IDs found. Exiting.")

        return  # Exit if there are no account IDs



    # AWS Organizations client

    session = create_aws_session()

    org_client = session.client('organizations')



    # Dictionary to hold account IDs and their corresponding tags

    account_tag_map = {}



    # Iterate through loaded account IDs and list their tags

    for account_id in account_ids:

        tags = list_tags(str(account_id), org_client)

        account_tag_map[account_id] = tags

       

        if tags is not None:

            logging.info(f'Tags for account {account_id}: {tags}')

        def main():
            parser = argparse.ArgumentParser(description='Fetch AWS account tags and write to Excel.')
            parser.add_argument('--excel_file', type=str, default='C:/Finops/AWSFinops/AWS_Account_Ids.xlsx', help='Path to input Excel file with AWS Account IDs')
            parser.add_argument('--output_file', type=str, default='AWS_Account_Tags_New.xlsx', help='Name for the output Excel file')
            args = parser.parse_args()

            excel_file = args.excel_file
            output_file = args.output_file
            account_ids = load_excel_file(excel_file)

            if not account_ids:
                logging.info("No account IDs found. Exiting.")
                return  # Exit if there are no account IDs

            # AWS Organizations client
            session = create_aws_session()
            org_client = session.client('organizations')

            # Dictionary to hold account IDs and their corresponding tags
            account_tag_map = {}

            # Iterate through loaded account IDs and list their tags
            for account_id in account_ids:
                tags = list_tags(str(account_id), org_client)
                account_tag_map[account_id] = tags
                if tags is not None:
                    logging.info(f'Tags for account {account_id}: {tags}')
                else:
                    logging.info(f'No tags found for account {account_id} or an error occurred.')

            # Write the tags to a new Excel file
            write_tags_to_excel(output_file, account_tag_map)

