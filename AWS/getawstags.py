import os
from typing import List, Tuple, Optional
import boto3
import openpyxl
import logging
import argparse
import pandas as pd
from dotenv import load_dotenv

# Load variables from .env
load_dotenv()

# Setup logging configuration
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# Load AWS credentials
#id = credentials.AWS_ACCESS_KEY_ID
id = os.getenv("AWS_ACCESS_KEY_ID")
key = os.getenv("AWS_SECRET_ACCESS_KEY")
region = os.getenv("AWS_DEFAULT_REGION")
def create_aws_session():
    """
    Create an AWS session using credentials loaded from the credentials module.
    """
    # Create an STS client using permanent credentials
    sts_client = boto3.client(
        'sts',
        aws_access_key_id=id,
        aws_secret_access_key=key,
        region_name=region

    )
  
    # Get temporary security credentials
    response = sts_client.get_session_token(DurationSeconds=3600)
    # Access temporary credentials from the response
    temp_credentials = response['Credentials']
    return boto3.Session(
        aws_access_key_id=temp_credentials['AccessKeyId'],
        aws_secret_access_key=temp_credentials['SecretAccessKey'],
        aws_session_token=temp_credentials['SessionToken'],
        region_name=region
    )
def load_excel_file(file_path):
    """
    Load the Excel file and return account IDs.
    :param file_path: The path to the Excel file
    :return: List of account IDs
    """
    account_ids = []
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming account IDs start from the second row
            account_id = row[0]  # Assuming the account ID is in the first column
            if account_id is not None:
                account_ids.append(account_id)  # Retain original format
    except Exception as e:
        logging.error(f"Failed to load Excel file {file_path}: {e}")
    return account_ids
def list_tags(account_id, org_client):
    """
    List tags for a given AWS Organizations account.
    :param account_id: The ID of the AWS account
    :param org_client: AWS Organizations client
    :return: Tags associated with the account
    """
    try:
        response = org_client.list_tags_for_resource(
            ResourceId=account_id
        )
        return response['Tags']
    except Exception as e:
        logging.error(f'Error retrieving tags for account {account_id}: {str(e)}')
        return None
def write_tags_to_excel(output_file, account_tag_map):
    """
    Write account IDs and their corresponding tags to a new Excel file.
    :param output_file: The path for the output Excel file

    :param account_tag_map: A dictionary mapping account IDs to their tags
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # Write headers
    headers = ['Account ID']
    # Create a set to hold unique tag keys
    unique_keys = set()
    for tags in account_tag_map.values():
        if tags:
            for tag in tags:
                unique_keys.add(tag['Key'])
    headers += list(unique_keys)

    sheet.append(headers)

    # Write account IDs and tags

    for account_id, tags in account_tag_map.items():

        row = [account_id]  # Account ID as is


        tag_dict = {tag['Key']: tag['Value'] for tag in tags} if tags else {}


        # Add values for each unique key; if the key is not present, add an empty string

        for key in unique_keys:

            row.append(tag_dict.get(key, ''))


        sheet.append(row)

    # Save the workbook


    try:


        workbook.save(output_file)
        logging.info(f"Tags successfully written to {output_file}.")
    except Exception as e:
        logging.error(f"Failed to write tags to Excel file {output_file}: {e}")
def load_excel(file_name: str):
    """
    Loads an Excel file from the root folder.

    Parameters:
        file_name (str): The name of the Excel file (e.g., 'example.xlsx').

    Returns:
        pandas.DataFrame: DataFrame containing the Excel file data.
    """
    root_folder = os.getcwd()
    file_path = os.path.join(root_folder, file_name)

    try:
        df = pd.read_excel(file_path)
        print(f"{file_name} loaded successfully from {root_folder}")
        return df
    except FileNotFoundError:
        print(f"Error: {file_name} not found in {root_folder}")
        return None


def load_account_ids(
    file_name_or_path: str,
    *,
    column: str = "AccountID",
    sheet_name: Optional[str] = None,
    search_subfolders: bool = True
) -> Tuple[List[str], pd.DataFrame]:
    """
    Load AWS account IDs from an Excel file.


    Args:
        file_name_or_path: Excel file name or absolute/relative path.
        column: Column in the sheet containing account IDs.
        sheet_name: Optional sheet name or index; if None, uses the first sheet.
        search_subfolders: If True and a bare file name is given, search cwd recursively.


    Returns:
        (account_ids, df): A list of account IDs and the loaded DataFrame.


    Raises:
        FileNotFoundError: If the file can't be located.
        ValueError: If the required column is missing or empty.
    """
    # Expand env vars and user (~)
    candidate = os.path.expandvars(os.path.expanduser(file_name_or_path))


    def _exists(p: str) -> bool:
        return os.path.isfile(p)


    # If it's a path and exists, use it
    if _exists(candidate):
        resolved = candidate
    else:
        # If only a filename was provided, try CWD and (optionally) walk subfolders
        base = os.path.basename(candidate)
        cwd_path = os.path.join(os.getcwd(), base)
        if _exists(cwd_path):
            resolved = cwd_path
        elif search_subfolders:
            resolved = None
            for dirpath, _, filenames in os.walk(os.getcwd()):
                if base in filenames:
                    resolved = os.path.join(dirpath, base)
                    break
            if not resolved:
                raise FileNotFoundError(
                    f"Excel file '{base}' not found in '{os.getcwd()}' or its subfolders."
                )
        else:
            raise FileNotFoundError(f"Excel file '{candidate}' not found.")


    # Read the Excel
    df = pd.read_excel(resolved, sheet_name=sheet_name)


    # If a specific sheet was selected, df is a DataFrame; if not and Excel has multiple sheets,
    # pandas may return a Dict[str, DataFrame]. Normalize to a DataFrame.
    if isinstance(df, dict):
        # Use the first sheet if none specified
        first_key = next(iter(df))
        df = df[first_key]


    # Validate column
    if column not in df.columns:
        raise ValueError(
            f"Required column '{column}' not found. Available columns: {list(df.columns)}"
        )


    # Drop NA and coerce to str, strip whitespace
    ids_series = (
        df[column]
        .dropna()
        .astype(str)
        .str.strip()
        .replace({"": pd.NA})
        .dropna()
    )


    if ids_series.empty:
        raise ValueError(f"No non-empty values found in column '{column}'.")


    account_ids = ids_series.tolist()
    return account_ids, df

def main():
 
    excel_path = "C:\\Users\\H630384\\project\\corp_enit_finops_tagging_governance\\AWS\\aws_accountid.xlsx"
    output_file = "AWS_Account_Tags_New.xlsx"  # Output file name
    # Call the function
    try:
        account_ids, df = load_account_ids(
            file_name_or_path=excel_path,
            column="account_ids",   # adjust if your column is named differently
            sheet_name=None,      # or "Sheet1"
            search_subfolders=True
        )


        # Use the IDs
        print(f"Loaded {len(account_ids)} AWS account IDs")
        print(account_ids[:10])  # preview first 10 IDs


        if not account_ids:

         logging.info("No account IDs found. Exiting.")

         return  
        # # Exit if there are no account IDs



        # AWS Organizations client

        session = create_aws_session()  
        org_client = session.client('organizations')
        # # Dictionary to hold account IDs and their corresponding tags

        account_tag_map = {}
          # # Iterate through loaded account IDs and list their tags

        for account_id in account_ids:
  
         tags = list_tags(str(account_id), org_client)

        account_tag_map[account_id] = tags

        if tags is not None:

            logging.info(f'Tags for account {account_id}: {tags}')

        else:

            logging.info(f'No tags found for account {account_id} or an error occurred.')
         
        # Write the tags to a new Excel file

        write_tags_to_excel(output_file, account_tag_map)

   

    except (FileNotFoundError, ValueError) as e:
        print(f" Error: {e}")

if __name__ == "__main__":

    main()

